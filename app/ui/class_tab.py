"""학급 로스터 출력 위젯(엑셀 스타일·동적 열·체크박스 열 없음).

행 선택은 엑셀처럼 한다 — 행을 클릭(단일), Ctrl+클릭(개별 토글),
Shift+클릭(범위). 선택된 행은 정상(무색), 선택 안 된 행은 흐리게 표시된다.
좌상단 코너의 체크박스로 전체 선택/해제. '체크된 행 생성'은 선택된 행을 채운다.

열: 학번 · 이름 + 동적 '내용' 열들.
엑셀식: 헤더 더블클릭=이름변경 / 헤더 우클릭=열 삽입·삭제 /
행번호 우클릭=행 삽입·삭제·전체선택 / 마지막 행 입력 시 자동 확장 / 클립보드 복붙.
"""
from __future__ import annotations

import csv
import json

from PySide6.QtCore import (QEvent, QItemSelection, QItemSelectionModel, QRectF, Qt,
                            Signal)
from PySide6.QtGui import QBrush, QColor, QKeySequence, QPalette, QTextDocument
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QCheckBox, QComboBox, QFileDialog,
    QHBoxLayout, QHeaderView, QInputDialog, QLabel, QMenu, QMessageBox,
    QPushButton, QStyle, QStyledItemDelegate, QTableWidget, QTableWidgetItem,
    QTabWidget, QVBoxLayout, QWidget,
)

from .. import config, theme

FIXED = 2                       # 학번, 이름
FIXED_LABELS = ["학번", "이름"]
DEFAULT_ROWS = 50
_DIM_BG = QColor("#f5f5f5")     # 선택 안 된 행 배경(흐림)
_DIM_FG = QColor("#b3b3b3")     # 선택 안 된 행 글자(흐림)
SPELL_ROLE = Qt.ItemDataRole.UserRole + 7   # 맞춤법 교정 색상 html 저장


class _RowDimDelegate(QStyledItemDelegate):
    """선택된 행은 무색(강조 제거), 선택 안 된 행은 흐리게.
    맞춤법 교정 html(SPELL_ROLE)이 있으면 색상 글자로 렌더링."""

    def initStyleOption(self, option, index):  # noqa: N802
        super().initStyleOption(option, index)
        if option.state & QStyle.StateFlag.State_Selected:
            option.state &= ~QStyle.StateFlag.State_Selected   # 파란 강조 제거 → 무색
        else:
            option.backgroundBrush = QBrush(_DIM_BG)
            option.palette.setColor(QPalette.ColorRole.Text, _DIM_FG)

    def paint(self, painter, option, index):  # noqa: N802
        html = index.data(SPELL_ROLE)
        if not html:
            super().paint(painter, option, index)
            return
        self.initStyleOption(option, index)
        painter.save()
        bg = (option.backgroundBrush
              if option.backgroundBrush.style() != Qt.BrushStyle.NoBrush
              else option.palette.base())
        painter.fillRect(option.rect, bg)
        doc = QTextDocument()
        doc.setDefaultFont(option.font)
        doc.setTextWidth(option.rect.width() - 8)
        doc.setHtml(html)
        painter.translate(option.rect.left() + 4, option.rect.top() + 3)
        doc.drawContents(painter, QRectF(0, 0, option.rect.width() - 8,
                                         option.rect.height()))
        painter.restore()


class RosterGrid(QTableWidget):
    """학번·이름 + 동적 '내용' 열. 엑셀식 행 선택(Ctrl/Shift)."""

    columnsChanged = Signal()
    rejectRequested = Signal(str)      # 셀 우클릭 '이 표현 버리기'(부정 피드백)

    def __init__(self, content_headers: list[str] | None = None,
                 rows: int = DEFAULT_ROWS):
        content = content_headers or ["내용"]
        super().__init__(rows, FIXED + len(content))
        self._bulk = True
        # 실행취소(Ctrl+Z)/다시실행(Ctrl+Y) 상태
        self._undo: list = []
        self._redo: list = []
        self._restoring = False        # 되돌리기/다시실행 중
        self._loading = False          # 엑셀 가져오기·불러오기 중(기록 안 함)
        self._group = None             # 묶음 기록(채움·붙여넣기·삭제) 중이면 list
        self._cache: dict = {}         # (row,col)->텍스트(이전값 조회용)
        self.setHorizontalHeaderLabels(FIXED_LABELS + content)
        self.setWordWrap(True)
        self.verticalHeader().setDefaultSectionSize(28)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

        hh = self.horizontalHeader()
        hh.sectionDoubleClicked.connect(self._rename_header)
        hh.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        hh.customContextMenuRequested.connect(self._col_menu)
        vh = self.verticalHeader()
        vh.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        vh.customContextMenuRequested.connect(self._row_menu)

        self.setItemDelegate(_RowDimDelegate(self))

        # 좌상단 코너 = 전체 선택/해제 체크박스
        self.setCornerButtonEnabled(False)
        self._corner_cb = QCheckBox(self)
        self._corner_cb.setTristate(True)
        self._corner_cb.setToolTip("전체 선택/해제")
        self._corner_cb.clicked.connect(self._corner_clicked)

        self._apply_resize()
        self.itemChanged.connect(self._maybe_grow)
        self.selectionModel().selectionChanged.connect(self._on_sel_changed)
        # 열 너비를 바꾸면 줄바꿈이 달라지므로 행 높이도 다시 맞춤
        self.horizontalHeader().sectionResized.connect(lambda *a: self._fit_rows())
        self._bulk = False
        self._position_corner()
        self._update_corner()
        self._fit_rows()
        # Ctrl+휠 확대/축소는 viewport로 오는 휠 이벤트를 필터로 가로채야 잡힌다.
        self._base_px = 13                        # 테마 QSS 기본 글꼴(px)
        self._zoom_px = 13
        self._ctrl_down = False
        self.viewport().installEventFilter(self)

    # --- 선택(엑셀식) ---
    def selected_rows(self) -> list[int]:
        return sorted({i.row() for i in self.selectionModel().selectedRows()})

    def set_all_selected(self, on: bool):
        if on:
            self.selectAll()
        else:
            self.clearSelection()

    def _select_rows(self, rows: list[int]):
        if not rows:
            self.clearSelection()
            return
        sel = QItemSelection()
        last = self.columnCount() - 1
        mdl = self.model()
        for r in rows:
            sel.select(mdl.index(r, 0), mdl.index(r, last))
        self.selectionModel().select(
            sel, QItemSelectionModel.SelectionFlag.ClearAndSelect
            | QItemSelectionModel.SelectionFlag.Rows)

    def _on_sel_changed(self, *_):
        self.viewport().update()
        self._update_corner()

    # --- 코너 전체선택 체크박스 ---
    def _corner_clicked(self, *_):
        rows = len(self.selected_rows())
        if self.rowCount() and rows >= self.rowCount():
            self.clearSelection()
        else:
            self.selectAll()

    def _update_corner(self):
        n, total = len(self.selected_rows()), self.rowCount()
        cb = self._corner_cb
        cb.blockSignals(True)
        if total and n >= total:
            cb.setCheckState(Qt.CheckState.Checked)
        elif n:
            cb.setCheckState(Qt.CheckState.PartiallyChecked)
        else:
            cb.setCheckState(Qt.CheckState.Unchecked)
        cb.blockSignals(False)

    def _position_corner(self):
        w = self.verticalHeader().width()
        h = self.horizontalHeader().height()
        size = 16
        self._corner_cb.setGeometry(max(0, (w - size) // 2 + 1),
                                    max(0, (h - size) // 2), size, size)

    def resizeEvent(self, e):  # noqa: N802
        super().resizeEvent(e)
        self._position_corner()

    def showEvent(self, e):  # noqa: N802
        super().showEvent(e)
        self._position_corner()

    # --- 열/헤더 ---
    def _apply_resize(self):
        hh = self.horizontalHeader()
        hh.setStretchLastSection(False)
        for c in range(self.columnCount()):     # 모든 열 수동 조절 가능(Interactive)
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
        self.setColumnWidth(0, 72)               # 학번
        self.setColumnWidth(1, 84)               # 이름
        for c in range(FIXED, self.columnCount()):
            self.setColumnWidth(c, 440)          # 내용(넓게, 드래그로 조절)

    def _fit_rows(self):
        """셀 내용(자동 줄바꿈)에 맞춰 행 높이를 조절 — 문장이 다 보이게."""
        self.resizeRowsToContents()

    # --- Ctrl+휠 = 시트 확대/축소(엑셀식) ---------------------------------
    def eventFilter(self, obj, e):
        if obj is self.viewport() and e.type() == QEvent.Type.Wheel:
            import os
            import sys
            from PySide6.QtWidgets import QApplication
            ad = e.angleDelta()
            pd = e.pixelDelta()
            ctrl_flag = Qt.KeyboardModifier.ControlModifier
            # 이벤트 modifier + 앱 키보드 상태 + 직접 추적한 Ctrl 상태(Wayland 대비 삼중)
            ctrl = (bool(e.modifiers() & ctrl_flag)
                    or bool(QApplication.keyboardModifiers() & ctrl_flag)
                    or getattr(self, "_ctrl_down", False))
            if os.environ.get("SGB_WHEEL_DEBUG"):
                print(f"[wheel] evmods={e.modifiers()!r} appmods={QApplication.keyboardModifiers()!r} "
                      f"ctrl_down={getattr(self,'_ctrl_down',False)} ctrl={ctrl} "
                      f"angle={ad.y()} pixel={pd.y()}", file=sys.stderr, flush=True)
            if ctrl:
                d = ad.y() or ad.x() or pd.y() or pd.x()   # 마우스·터치패드 모두 대응
                self._zoom(1 if d >= 0 else -1)
                return True                        # 스크롤 대신 확대/축소로 소비
        return super().eventFilter(obj, e)

    def _reset_zoom(self) -> None:
        self._zoom_px = self._base_px
        self._apply_zoom_css()

    # --- 셀 우클릭: '이 표현 버리기'(부정 피드백) ---------------------------
    def contextMenuEvent(self, e):
        idx = self.indexAt(e.pos())
        if idx.isValid() and idx.column() >= FIXED:
            it = self.item(idx.row(), idx.column())
            if it and it.text().strip():
                m = QMenu(self)
                m.addAction(theme.ui_icon("trash", 16, "#c62828"),
                            "이 표현 버리기(다음부터 안 나오게)",
                            lambda: self._reject_cell(idx.row(), idx.column()))
                m.exec(e.globalPos())
                return
        super().contextMenuEvent(e)

    def _reject_cell(self, r: int, c: int) -> None:
        it = self.item(r, c)
        text = it.text().strip() if it else ""
        if not text:
            return
        self.rejectRequested.emit(text)            # 상위(AreaTab)에서 부정 예시로 저장
        self._begin_group()                        # 셀 비우기(되돌리기 가능)
        it.setText("")
        self._end_group()

    def _zoom(self, step: int) -> None:
        """표 글꼴 크기를 키우거나 줄인다. 앱 QSS의 `* {font-size}`가 setFont를 덮으므로
        위젯 자체 스타일시트로 글꼴 크기를 지정해야 실제로 반영된다."""
        self._zoom_px = max(9, min(40, self._zoom_px + step))
        self._apply_zoom_css()

    def _apply_zoom_css(self) -> None:
        px = self._zoom_px
        self.setStyleSheet(
            f"QTableWidget {{ font-size: {px}px; }}"
            f" QHeaderView::section {{ font-size: {px}px; }}")
        self._fit_rows()
        self._position_corner()

    def content_headers(self) -> list[str]:
        out = []
        for c in range(FIXED, self.columnCount()):
            it = self.horizontalHeaderItem(c)
            out.append(it.text() if it else f"내용{c - FIXED + 1}")
        return out

    def all_headers(self) -> list[str]:
        out = []
        for c in range(self.columnCount()):
            it = self.horizontalHeaderItem(c)
            out.append(it.text() if it else str(c))
        return out

    def _rename_header(self, idx: int):
        cur = self.horizontalHeaderItem(idx)
        cur = cur.text() if cur else ""
        name, ok = QInputDialog.getText(self, "열 이름 변경", "새 이름:", text=cur)
        if ok and name.strip():
            self.setHorizontalHeaderItem(idx, QTableWidgetItem(name.strip()))
            self.columnsChanged.emit()

    def _col_menu(self, pos):
        col = self.horizontalHeader().logicalIndexAt(pos)
        if col < 0:
            return
        m = QMenu(self)
        m.addAction("이름 변경", lambda: self._rename_header(col))
        m.addAction("◀ 왼쪽에 열 삽입", lambda: self._insert_col(max(FIXED, col)))
        m.addAction("오른쪽에 열 삽입 ▶", lambda: self._insert_col(max(FIXED, col + 1)))
        act = m.addAction("열 삭제", lambda: self.remove_content_column(col))
        act.setEnabled(col >= FIXED and self.columnCount() - FIXED > 1)
        m.exec(self.horizontalHeader().mapToGlobal(pos))

    def _row_menu(self, pos):
        row = self.verticalHeader().logicalIndexAt(pos)
        m = QMenu(self)
        if row >= 0:
            m.addAction("▲ 위에 행 삽입", lambda: self.insertRow(row))
            m.addAction("아래에 행 삽입 ▼", lambda: self.insertRow(row + 1))
            m.addAction("행 삭제", lambda: self.removeRow(row))
            m.addSeparator()
        m.addAction("전체 선택", lambda: self.selectAll())
        m.addAction("전체 해제", lambda: self.clearSelection())
        m.exec(self.verticalHeader().mapToGlobal(pos))

    def _insert_col(self, idx: int):
        self.insertColumn(idx)
        self.setHorizontalHeaderItem(
            idx, QTableWidgetItem(f"내용{self.columnCount() - FIXED}"))
        self._apply_resize()
        self.columnsChanged.emit()

    def add_content_column(self, name: str | None = None):
        self._insert_col(self.columnCount())
        if name:
            self.setHorizontalHeaderItem(self.columnCount() - 1,
                                         QTableWidgetItem(name))
            self.columnsChanged.emit()

    def remove_content_column(self, col: int) -> bool:
        if col < FIXED or self.columnCount() - FIXED <= 1:
            return False
        self.removeColumn(col)
        self._apply_resize()
        self.columnsChanged.emit()
        return True

    def _maybe_grow(self, item):
        r, c, new = item.row(), item.column(), item.text()
        old = self._cache.get((r, c), "")
        if old != new:
            if not (self._loading or self._restoring):
                self._record(r, c, old)          # 실행취소용 기록
            self._cache[(r, c)] = new
        if self._bulk or self._loading or self._restoring:
            return
        self.resizeRowToContents(r)              # 편집 시 그 행 높이 자동 맞춤
        if item.data(SPELL_ROLE):                # 수정하면 교정 색상 표시 해제
            self._bulk = True
            item.setData(SPELL_ROLE, None)
            self._bulk = False
        if r >= self.rowCount() - 1 and new.strip():
            self._bulk = True
            self.setRowCount(self.rowCount() + 10)
            self._bulk = False
            self._position_corner()

    # --- 실행취소(Undo)/다시실행(Redo) ---
    def _record(self, r, c, old):
        if self._group is not None:
            self._group.append((r, c, old))
        else:
            self._undo.append([(r, c, old)])
            if len(self._undo) > 300:
                self._undo.pop(0)
        self._redo.clear()

    def _begin_group(self):
        self._group = []

    def _end_group(self):
        if self._group:
            self._undo.append(self._group)
            if len(self._undo) > 300:
                self._undo.pop(0)
            self._redo.clear()
        self._group = None

    def _set_cell(self, r, c, text):
        if r >= self.rowCount():
            self.setRowCount(r + 1)
        self.setItem(r, c, QTableWidgetItem(text))
        self._cache[(r, c)] = text

    def _apply(self, group, stack):
        rev = []
        self._restoring = True
        for r, c, val in reversed(group):
            it = self.item(r, c)
            rev.append((r, c, it.text() if it else ""))
            self._set_cell(r, c, val)
        self._restoring = False
        stack.append(list(reversed(rev)))
        self._fit_rows()
        self.viewport().update()

    def undo(self):
        if self._undo:
            self._apply(self._undo.pop(), self._redo)

    def redo(self):
        if self._redo:
            self._apply(self._redo.pop(), self._undo)

    def _rebuild_cache(self):
        self._cache = {}
        for r in range(self.rowCount()):
            for c in range(self.columnCount()):
                it = self.item(r, c)
                if it and it.text():
                    self._cache[(r, c)] = it.text()

    # --- 클립보드 ---
    def keyReleaseEvent(self, e):  # noqa: N802
        if e.key() == Qt.Key.Key_Control:
            self._ctrl_down = False
        super().keyReleaseEvent(e)

    def keyPressEvent(self, e):  # noqa: N802
        if e.key() == Qt.Key.Key_Control:
            self._ctrl_down = True
        if e.matches(QKeySequence.StandardKey.Undo):
            self.undo()
        elif e.matches(QKeySequence.StandardKey.Redo):
            self.redo()
        elif e.matches(QKeySequence.StandardKey.Copy):
            self._copy()
        elif e.matches(QKeySequence.StandardKey.Paste):
            self._paste()
        elif e.matches(QKeySequence.StandardKey.Delete):
            self._begin_group()
            for it in self.selectedItems():
                it.setText("")
            self._end_group()
        elif (e.modifiers() & Qt.KeyboardModifier.ControlModifier
              and e.key() in (Qt.Key.Key_Plus, Qt.Key.Key_Equal)):
            self._zoom(1)                          # Ctrl + '+' 확대
        elif (e.modifiers() & Qt.KeyboardModifier.ControlModifier
              and e.key() == Qt.Key.Key_Minus):
            self._zoom(-1)                         # Ctrl + '-' 축소
        elif (e.modifiers() & Qt.KeyboardModifier.ControlModifier
              and e.key() == Qt.Key.Key_0):
            self._reset_zoom()                     # Ctrl + '0' 원래대로
        else:
            super().keyPressEvent(e)

    def _copy(self):
        rngs = self.selectedRanges()
        if not rngs:
            return
        r = rngs[0]
        lines = []
        for row in range(r.topRow(), r.bottomRow() + 1):
            cells = [self.item(row, c).text() if self.item(row, c) else ""
                     for c in range(r.leftColumn(), r.rightColumn() + 1)]
            lines.append("\t".join(cells))
        QApplication.clipboard().setText("\n".join(lines))

    def _paste(self):
        text = QApplication.clipboard().text()
        if not text:
            return
        rows = text.replace("\r\n", "\n").replace("\r", "\n").rstrip("\n").split("\n")
        start_r = max(0, self.currentRow())
        start_c = max(0, self.currentColumn())
        self._begin_group()
        self._bulk = True
        if start_r + len(rows) > self.rowCount():
            self.setRowCount(start_r + len(rows))
        for i, line in enumerate(rows):
            for j, val in enumerate(line.split("\t")):
                col = start_c + j
                if col >= self.columnCount():
                    break
                self.setItem(start_r + i, col, QTableWidgetItem(val))
        self._bulk = False
        self._end_group()
        self._fit_rows()

    # --- 채움 ---
    def fill_selected(self, variants: list[str], col: int) -> int:
        rows = self.selected_rows()
        self._begin_group()
        self._bulk = True
        for k, r in enumerate(rows):
            if k >= len(variants):
                break
            self.setItem(r, col, QTableWidgetItem(variants[k]))
        self._bulk = False
        self._end_group()
        self._fit_rows()
        return min(len(rows), len(variants))

    def column_contents(self, col: int) -> list[str]:
        out = []
        for r in range(self.rowCount()):
            it = self.item(r, col)
            if it and it.text().strip():
                out.append(it.text().strip())
        return out

    # --- 데이터 ---
    def export_rows(self) -> list[list[str]]:
        headers = [self.horizontalHeaderItem(c).text()
                   if self.horizontalHeaderItem(c) else ""
                   for c in range(self.columnCount())]
        out = [headers]
        for r in range(self.rowCount()):
            vals = [self.item(r, c).text().strip() if self.item(r, c) else ""
                    for c in range(self.columnCount())]
            if any(vals):
                out.append(vals)
        return out

    def import_rows(self, table: list[list[str]]):
        if not table:
            return
        first = table[0]
        has_header = not (first and (first[0] or "").strip().isdigit())
        if has_header:
            cols = list(first)
            data = table[1:]
        else:
            cols = [self.horizontalHeaderItem(c).text() if self.horizontalHeaderItem(c)
                    else "" for c in range(self.columnCount())]
            data = table
        content = cols[FIXED:] or ["내용"]
        self._loading = True
        self._bulk = True
        self.setColumnCount(FIXED + len(content))
        self.setHorizontalHeaderLabels(FIXED_LABELS + content)
        self._apply_resize()
        self.setRowCount(max(DEFAULT_ROWS, len(data)))
        filled = []
        for r, row in enumerate(data):
            for j, val in enumerate(row):
                if j < self.columnCount():
                    self.setItem(r, j, QTableWidgetItem(str(val)))
            if any((c or "").strip() for c in row):
                filled.append(r)
        self._bulk = False
        self._loading = False
        self._rebuild_cache()
        self._undo.clear()
        self._redo.clear()
        self._select_rows(filled)          # 명단 있는 행 자동 선택
        self._position_corner()
        self._fit_rows()
        self.columnsChanged.emit()

    def to_dict(self) -> dict:
        return {"headers": self.content_headers(),
                "rows": self.export_rows()[1:]}

    def load_dict(self, d):
        if isinstance(d, list):
            d = {"headers": ["내용"], "rows": d}
        content = d.get("headers") or ["내용"]
        rows = d.get("rows") or []
        self._loading = True
        self._bulk = True
        self.setColumnCount(FIXED + len(content))
        self.setHorizontalHeaderLabels(FIXED_LABELS + content)
        self._apply_resize()
        self.setRowCount(max(DEFAULT_ROWS, len(rows)))
        filled = []
        for r, row in enumerate(rows):
            for j, val in enumerate(row):
                if j < self.columnCount():
                    self.setItem(r, j, QTableWidgetItem(str(val)))
            if any((c or "").strip() for c in row):
                filled.append(r)
        self._bulk = False
        self._loading = False
        self._rebuild_cache()
        self._undo.clear()
        self._redo.clear()
        self._select_rows(filled)
        self._fit_rows()


class TabbedRoster(QWidget):
    """학급 시트 탭(엑셀식) + 동적 열 + 생성 대상 열 선택."""

    saveRequested = Signal()        # '저장'(내용 저장 + 자동 학습)
    fullscreenRequested = Signal()  # 시트 전체화면 토글
    rejectRequested = Signal(str)   # 셀 '버리기'(부정 피드백) → AreaTab로 전달

    def __init__(self, store_key: str):
        super().__init__()
        self._store = config.DATA_DIR / f"roster_{store_key}.json"
        v = QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)

        bar = QHBoxLayout()
        bar.setContentsMargins(0, 2, 0, 2)
        bar.addWidget(QLabel("생성 대상 열"))
        self.col_combo = QComboBox()
        self.col_combo.setToolTip("'체크된 행 생성' 결과가 들어갈 열을 고릅니다.")
        self.col_combo.setMinimumWidth(120)
        bar.addWidget(self.col_combo)
        bar.addSpacing(16)
        hint = QLabel("행 클릭=선택 · Ctrl+클릭=개별 · Shift+클릭=범위 · 좌상단□=전체 · "
                      "헤더 더블클릭=이름변경 · 우클릭=열/행 · Ctrl+휠 또는 Ctrl+±=확대/축소")
        hint.setStyleSheet("color:#999;")
        bar.addWidget(hint)
        bar.addStretch(1)
        chk = QPushButton(" 맞춤법 검사")
        chk.setIcon(theme.ui_icon("check", 16))
        chk.setToolTip("맞춤법·띄어쓰기를 교정합니다.\n"
                       "정확한 검사는 네이버(온라인) 전송이 필요하며, 처음 누를 때 동의를 "
                       "묻습니다. 동의하지 않으면 오프라인 검사만 수행합니다.")
        chk.clicked.connect(self._spellcheck)
        imp = QPushButton(" 엑셀 가져오기")
        imp.setIcon(theme.ui_icon("import", 16))
        imp.clicked.connect(self._import)
        exp = QPushButton(" 엑셀 내보내기")
        exp.setIcon(theme.ui_icon("export", 16))
        exp.clicked.connect(self._export)
        save = QPushButton(" 저장")
        save.setIcon(theme.ui_icon("save", 16, "#ffffff"))   # 초록 버튼 위 → 흰색
        save.setToolTip("표 내용을 저장하고, 저장과 동시에 자동으로 학습합니다.")
        save.setObjectName("saveBtn")
        save.clicked.connect(self.saveRequested.emit)
        bar.addWidget(chk)
        bar.addWidget(imp)
        bar.addWidget(exp)
        bar.addWidget(save)
        v.addLayout(bar)

        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        # 전체화면 버튼은 탭바 우측 코너에(＋ 학급은 마지막 탭 뒤 '＋' 탭 = 엑셀식)
        full = QPushButton(" 전체화면")
        full.setIcon(theme.ui_icon("maximize", 16))
        full.setToolTip("시트를 전체화면으로 크게 봅니다. (Esc로 해제)")
        full.setObjectName("fullBtn")
        full.clicked.connect(self.fullscreenRequested.emit)
        self.tabs.setCornerWidget(full, Qt.Corner.TopRightCorner)
        self.tabs.tabBarDoubleClicked.connect(self._rename_class)
        self.tabs.tabBarClicked.connect(self._on_tab_clicked)   # '＋' 탭 클릭 → 학급 추가
        tb = self.tabs.tabBar()
        tb.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        tb.customContextMenuRequested.connect(self._tab_menu)
        self.tabs.currentChanged.connect(self._refresh_cols)
        v.addWidget(self.tabs, 1)

        self._plus_page = None
        self._load()
        if self._class_count() == 0:
            self._new_grid_tab("1반")
        self._ensure_plus_tab()          # 엑셀식: 마지막 학급 탭 바로 뒤에 '＋'
        self.tabs.setCurrentIndex(0)
        self._refresh_cols()

    def _new_grid_tab(self, name: str, grid: RosterGrid | None = None) -> int:
        g = grid or RosterGrid()
        g.columnsChanged.connect(self._refresh_cols)
        g.rejectRequested.connect(self.rejectRequested)      # 셀 '버리기' 전달
        return self.tabs.addTab(g, name)

    # --- 엑셀식 '＋' 탭(마지막 학급 탭 뒤) ------------------------------------
    def _ensure_plus_tab(self) -> int:
        """'＋' 탭을 항상 맨 끝에 둔다."""
        if self._plus_page is not None:
            idx = self.tabs.indexOf(self._plus_page)
            if idx != -1:
                self.tabs.removeTab(idx)
        self._plus_page = QWidget()
        i = self.tabs.addTab(self._plus_page, "＋")
        self.tabs.setTabToolTip(i, "학급(시트) 추가")
        return i

    def _is_plus(self, i: int) -> bool:
        return 0 <= i < self.tabs.count() and self.tabs.widget(i) is self._plus_page

    def _class_count(self) -> int:
        return sum(1 for i in range(self.tabs.count())
                   if isinstance(self.tabs.widget(i), RosterGrid))

    def _on_tab_clicked(self, i: int) -> None:
        if self._is_plus(i):
            self._add_class()

    def active(self) -> RosterGrid | None:
        w = self.tabs.currentWidget()
        return w if isinstance(w, RosterGrid) else None

    def _refresh_cols(self):
        g = self.active()
        if not g:
            return
        prev = self.col_combo.currentText()
        self.col_combo.blockSignals(True)
        self.col_combo.clear()
        self.col_combo.addItems(g.all_headers())          # 모든 열(학번·이름 포함)
        i = self.col_combo.findText(prev)
        # 기본 대상 = 첫 내용 열(학번·이름이 아니라 '내용')
        self.col_combo.setCurrentIndex(i if i >= 0 else min(FIXED, self.col_combo.count() - 1))
        self.col_combo.blockSignals(False)

    def _target_col(self) -> int:
        idx = self.col_combo.currentIndex()
        return max(0, idx)

    def checked_count(self) -> int:
        g = self.active()
        return len(g.selected_rows()) if g else 0

    def fill_checked(self, variants: list[str]) -> int:
        g = self.active()
        return g.fill_selected(variants, self._target_col()) if g else 0

    def contents(self) -> list[str]:
        g = self.active()
        return g.column_contents(self._target_col()) if g else []

    # --- 학급 시트 탭 ---
    def _add_class(self):
        default = f"{self._class_count() + 1}반"
        name, ok = QInputDialog.getText(self, "학급 추가", "학급 이름:", text=default)
        if not (ok and name.strip()):
            self.tabs.setCurrentIndex(0)          # '＋'에 머무르지 않도록 복귀
            return
        g = RosterGrid()
        g.columnsChanged.connect(self._refresh_cols)
        g.rejectRequested.connect(self.rejectRequested)
        idx = self.tabs.indexOf(self._plus_page)  # '＋' 바로 앞에 삽입
        self.tabs.insertTab(idx, g, name.strip())
        self.tabs.setCurrentIndex(idx)

    def _rename_class(self, i: int):
        if i < 0 or self._is_plus(i):
            return
        name, ok = QInputDialog.getText(self, "이름 변경", "새 학급 이름:",
                                        text=self.tabs.tabText(i))
        if ok and name.strip():
            self.tabs.setTabText(i, name.strip())

    def _tab_menu(self, pos):
        i = self.tabs.tabBar().tabAt(pos)
        if i < 0 or self._is_plus(i):
            return
        m = QMenu(self)
        m.addAction("이름 변경", lambda: self._rename_class(i))
        m.addAction("학급 삭제", lambda: self._remove_class(i))
        m.exec(self.tabs.tabBar().mapToGlobal(pos))

    def _remove_class(self, i: int):
        if self._is_plus(i):
            return
        if self._class_count() == 1:
            QMessageBox.information(self, "삭제 불가", "학급은 최소 1개 있어야 합니다.")
            return
        if QMessageBox.question(
                self, "학급 삭제",
                f"'{self.tabs.tabText(i)}' 학급을 삭제할까요? 표 내용도 사라집니다.") == \
                QMessageBox.StandardButton.Yes:
            self.tabs.removeTab(i)
            if self._is_plus(self.tabs.currentIndex()):   # 삭제 후 '＋'에 남지 않게
                self.tabs.setCurrentIndex(0)

    # --- 맞춤법 검사(네이버 온라인 + 오프라인 폴백) ---
    def _spellcheck(self):
        from PySide6.QtGui import QCursor, QGuiApplication
        from .. import spellcheck, settings
        g = self.active()
        if not g:
            return
        cells = [(r, c, g.item(r, c).text().strip())
                 for r in range(g.rowCount())
                 for c in range(FIXED, g.columnCount())
                 if g.item(r, c) and g.item(r, c).text().strip()]
        if not cells:
            QMessageBox.information(self, "맞춤법 검사", "검사할 내용이 없습니다.")
            return
        # 온라인 전송 동의(최초 1회) — 기본은 전송하지 않고 오프라인 검사
        if not settings.get("spell_online_consent", False):
            ret = QMessageBox.question(
                self, "맞춤법 검사 — 온라인 전송 안내",
                "맞춤법 검사는 표 '내용'의 문장을 네이버 맞춤법 검사기(온라인)로 "
                "전송합니다.\n이름·학번은 보내지 않지만, 평가 문장 자체는 인터넷을 "
                "거칩니다.\n\n"
                "• 예: 네이버로 전송해 정확히 검사(이후 다시 묻지 않음)\n"
                "• 아니오: 전송하지 않고 오프라인 검사만 수행\n\n"
                "네이버로 전송할까요?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No)
            if ret == QMessageBox.StandardButton.Yes:
                settings.set("spell_online_consent", True)
            else:
                self._spellcheck_offline(g)          # 전송 없이 오프라인 검사
                return
        QGuiApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))
        online, applied = True, []
        try:
            for (r, c, t) in cells:
                res = spellcheck.naver_spellcheck(t)
                if res is None:
                    online = False
                    break
                corrected, n, html = res
                if n > 0 and corrected and corrected != t:
                    applied.append((r, c, corrected, spellcheck.styled_html(html)))
        finally:
            QGuiApplication.restoreOverrideCursor()

        if not online:
            self._spellcheck_offline(g)
            return
        if not applied:
            QMessageBox.information(self, "맞춤법 검사", "맞춤법 오류가 없습니다.")
            return
        # 교정문 적용 + 셀에 색상(빨강=맞춤법, 파랑=띄어쓰기) 표시
        g._begin_group()
        g._bulk = True
        for (r, c, corrected, html) in applied:
            it = QTableWidgetItem(corrected)
            it.setData(SPELL_ROLE, html)
            g.setItem(r, c, it)
        g._bulk = False
        g._end_group()
        g._fit_rows()
        g.viewport().update()
        g._select_rows(sorted({r for (r, c, _, _) in applied}))
        QMessageBox.information(
            self, "맞춤법 검사",
            f"{len(applied)}곳을 교정해 색상으로 표시했습니다.\n"
            "빨강=맞춤법, 파랑=띄어쓰기. 셀을 수정하면 색 표시가 사라집니다.")

    def _spellcheck_offline(self, g):
        from ..spellcheck import suspect_words
        flagged, lines = [], []
        for r in range(g.rowCount()):
            words = []
            for c in range(FIXED, g.columnCount()):
                it = g.item(r, c)
                if it and it.text().strip():
                    words += suspect_words(it.text())
            if words:
                flagged.append(r)
                name = g.item(r, 1).text() if g.item(r, 1) else ""
                lines.append(f"{r + 1}행 {name}: {', '.join(dict.fromkeys(words))}")
        if flagged:
            g._select_rows(flagged)
            QMessageBox.warning(
                self, "맞춤법 검사(오프라인)",
                "온라인 검사에 연결하지 못해 오프라인으로 의심 단어만 표시합니다:\n\n"
                + "\n".join(lines[:40]))
        else:
            QMessageBox.information(self, "맞춤법 검사", "의심되는 단어가 없습니다.")

    # --- 엑셀 입출력 ---
    def _import(self):
        g = self.active()
        if not g:
            return
        from pathlib import Path
        start = str(Path.home() / "Downloads")
        if not Path(start).exists():
            start = str(Path.home())
        path, _ = QFileDialog.getOpenFileName(
            self, "엑셀 가져오기", start, "엑셀/CSV (*.xlsx *.csv);;모든 파일 (*)",
            options=QFileDialog.Option.DontUseNativeDialog)
        if not path:
            return
        try:
            data = _read_table(path)
        except Exception as e:  # noqa: BLE001
            QMessageBox.warning(self, "가져오기 실패", f"파일을 읽지 못했습니다:\n{e}")
            return
        g.import_rows(data)
        self._refresh_cols()

    def _export(self):
        g = self.active()
        if not g:
            return
        from pathlib import Path
        path, _ = QFileDialog.getSaveFileName(
            self, "엑셀 내보내기",
            str(Path.home() / "Downloads" / f"{self.tabs.tabText(self.tabs.currentIndex())}.xlsx"),
            "엑셀 (*.xlsx);;CSV (*.csv)",
            options=QFileDialog.Option.DontUseNativeDialog)
        if not path:
            return
        rows = g.export_rows()
        try:
            _write_table(path, rows)
        except Exception as e:  # noqa: BLE001
            QMessageBox.warning(self, "내보내기 실패", f"저장하지 못했습니다:\n{e}")
            return
        QMessageBox.information(self, "완료", f"{len(rows) - 1}행을 저장했습니다.")

    # --- 영속 저장 ---
    def save(self) -> None:
        data = {self.tabs.tabText(i): self.tabs.widget(i).to_dict()
                for i in range(self.tabs.count())
                if isinstance(self.tabs.widget(i), RosterGrid)}   # '＋' 탭 제외
        try:
            self._store.write_text(json.dumps(data, ensure_ascii=False),
                                   encoding="utf-8")
        except OSError:
            pass

    def _load(self) -> None:
        if not self._store.exists():
            return
        try:
            data = json.loads(self._store.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return
        for name, d in data.items():
            g = RosterGrid()
            g.load_dict(d)
            self._new_grid_tab(name, g)


def _read_table(path: str) -> list[list[str]]:
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        out = [["" if v is None else str(v) for v in row]
               for row in ws.iter_rows(values_only=True)]
        wb.close()
        return [r for r in out if any(c.strip() for c in r)]
    with open(path, encoding="utf-8-sig", newline="") as f:
        return [row for row in csv.reader(f) if any(c.strip() for c in row)]


def _write_table(path: str, rows: list[list[str]]) -> None:
    if path.lower().endswith(".csv"):
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(rows)
        return
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "생기부"
    for r in rows:
        ws.append(r)
    ncol = len(rows[0]) if rows else 0
    for ci in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12 if ci <= 2 else 60
    for row in ws.iter_rows():
        for ci, cell in enumerate(row, start=1):
            if ci >= 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
